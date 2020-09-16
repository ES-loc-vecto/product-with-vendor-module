# -*- coding: utf-8 -*-

import time
import pytz
import json
import datetime
import io
import logging
import base64
import re
from datetime import date, datetime
from openerp import tools
from odoo import api, fields, models, _
from odoo.tools import date_utils
from odoo.exceptions import UserError, ValidationError

from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

_logger = logging.getLogger(__name__)


class VendorEvaluation(models.Model):
    _name = 'circles.vendor.evaluation'
    _description = 'Evaluation list for product'

    name = fields.Char('Name', related='criteria_id.name')
    criteria_id = fields.Many2one('circles.vendor.criteria', required=True)
    attachment_ids = fields.Many2many('ir.attachment', string="Attachment")
    vendor_id = fields.Many2one('res.partner')
    data = fields.Char('Detail', defualt='')
    rating = fields.Float('Rating')
    is_verified = fields.Boolean(string='Verified?', default=True)


class VendorCriteria(models.Model):
    _name = 'circles.vendor.criteria'
    _description = 'Criteria list for product'

    name = fields.Char('Name', translate=True)
    description = fields.Char('Description', default='')
    is_required = fields.Boolean('Required')
    active = fields.Boolean('Active', default=True)


class JobPosition(models.Model):
    _name = 'circles.job.position'

    name = fields.Char('Job position')


class PartnerIndustries(models.Model):
    _inherit = 'res.partner.industry'

    vendor_ids = fields.Many2many('res.partner', string='Supplier')
    color = fields.Integer('Color Index')


class Vendor(models.Model):
    _inherit = 'res.partner'

    product_ids = fields.One2many('product.supplierinfo', 'name')
    vendor_evaluation_ids = fields.One2many('circles.vendor.evaluation', 'vendor_id', string='Evaluation')
    vendor_evaluation_search = fields.Many2one('circles.vendor.criteria', string='Have Criteria')
    num_empl = fields.Char('Employee number')
    is_geo_localize_updated = fields.Boolean('Is geo_localize updated', default=False)
    user_id = fields.Many2one('res.users', string='PIC', help='The internal user in charge of this contact.') # rename the original verson
    description = fields.Text(string='Vendor Description')
    function = fields.Many2one('circles.job.position', string='Job Position')

    industries_ids = fields.Many2many('res.partner.industry', 'vendor_ids', string='Categories') # it've the potential of confusion

    @api.model
    def _get_capacity_filter(self):
        return [('category_id', '=', self.env['uom.category'].search([('name','=', 'Working Time')]).id )]
    capacity_uom_id = fields.Many2one('uom.uom', 'Capacity unit', domain=_get_capacity_filter)
    capacity_current = fields.Float(string='Capacity', size=15, digits=(15, 0))
    capacity_compute =fields.Float(string='capacity', size=15, digits=(15, 0), compute='_compute_capacity_current', store=True)
    @api.depends('capacity_current')
    def _compute_capacity_current(self):
        for record in self:
            if record.capacity_current != 0:
                record.capacity_compute = record.capacity_current
            else:
                record.capacity_compute = None


    established_date = fields.Char('Established date')
    established_date_compute = fields.Date('established date', compute='_compute_established_date', store=True)
    @api.constrains('established_date')
    def _validate_established_date(self):
        for record in self:
            if record.established_date:
                if not re.search(r'(^\d+/\d+/\d+$)|(^\d+/\d+$)|(^\d+$)', record.established_date):
                    raise ValidationError("Wrong date format")
    def _try_parsing_date(self, text):
        for fmt in ('%d/%m/%Y', '%m/%Y', '%Y'):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
        return None
    @api.depends('established_date')
    def _compute_established_date(self):
        for record in self:
            if record.established_date:
                record.established_date_compute = self._try_parsing_date(record.established_date)
            else:
                record.established_date_compute = None

    type = fields.Selection(
        [('contact', 'Contact'),
         ('invoice', 'Invoice Address'),
         ('delivery', 'Delivery Address'),
         ('other', 'Other Address'),
         ("private", "Private Address"),
         ("production", "Production Address"),
        ], string='Address Type',
        default='contact',
        help="Invoice & Delivery addresses are used in sales orders. Private addresses are only visible by authorized users.")


    @api.onchange('established_date')
    def _validate_established_date_onchange(self):
        if self.established_date:
            if not re.search(r'(^\d+/\d+/\d+$)|(^\d+/\d+$)|(^\d+$)', self.established_date):
                return {
                    'warning': {
                        'title': 'Error',
                        'message': 'Wrong date format'

                    }
                }

    @api.onchange('name')
    def _format_name_onchange(self):
        if self.name:
            name_format = ' '.join([ a.title() if not a.isupper() else a for a in self.name.split()]) 
            self.name = name_format

    def gmap(self):
        ids = self.env.context.get('active_ids') # Get the ids from the current context

        # Replace office address with manufactory address
        for id in ids:
            for child_id in self.browse(id).child_ids:
                if child_id.type == 'production':
                    ids[ids.index(id)] = child_id.id

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'res.partner',
            'domain': [('id', 'in', ids)],
            'view_mode': 'map',
            'view_type': 'map',
            'target': 'new',
            }

    @api.onchange('street, state_id, country_id, city')
    def updated_geo_localize_flag(self):
        self.is_geo_localize_updated = False

    @api.model
    def update_geo_localize_vendors(self):
        list_vendors = self.search([('is_company', '=', True), ('supplier_rank', '=', 1), ('is_geo_localize_updated', '=', False)])

        count = 0
        for vendor in list_vendors:
            print('Id: %d, name: %s, progress %d/%d'%(vendor.id, vendor.name, count, len(list_vendors)))
            try:
                vendor.geo_localize()
                vendor.env.cr.execute("update res_partner set is_geo_localize_updated=True where id=%d"%(vendor.id))
                vendor.env.cr.commit()
            except Exception:
                self._send_togroup_notification("geo_localize got error at %s!"%(vendor.name))
                _logger.exception("Failed processing geo_localize")
            count += 1

    @api.model
    def _send_togroup_notification(self, msg):
        print("send notification!")
        self.env['mail.message'].create({
            'email_from': self.env.user.partner_id.email, #add the sender email
            'author_id': self.env.user.partner_id.id, # add the creator id
            'model': 'mail.channel', # model should be mail.channel
            'subtype_id': self.env.ref('mail.mt_comment').id, #Leave this as it is
            'body': msg, # here add the message body
            'channel_ids': [(4, self.env.ref('circles.channel_sys_notify_group').id)], # This is the channel where you want to send the message and all the users of this channel will receive message
            'res_id': self.env.ref('circles.channel_sys_notify_group').id, # here add the channel you created.
            'message_type': 'notification',
        })


class ProductofSupplierUserDefineFields(models.Model):
    _name = 'circles.supplierinfo.userdefine.fields'
    _description = 'fields for user define in supplierinfo model'

    name = fields.Char('Field Name')


class ProductofSupplierUserDefine(models.Model):
    _name = 'circles.supplierinfo.userdefine'
    _description = ' User define fields for supplierinfo model'

    supplierinfo_id = fields.Many2one('product.supplierinfo')
    field_define_id = fields.Many2one('circles.supplierinfo.userdefine.fields')
    field_data = fields.Char ('Value')
    field_unit_id = fields.Many2one('uom.uom', 'Unit (optional)')


class ProductofSupplier(models.Model):
    _name = 'product.supplierinfo'
    _inherit = ['product.supplierinfo','mail.thread']

    capacity_current = fields.Integer('Current capacity')
    capacity_maximum = fields.Integer('Maximum capacity')

    @api.model
    def _get_capacity_filter(self):
        return [('category_id', '=', self.env['uom.category'].search([('name','=', 'Working Time')]).id )]
    capacity_uom_id = fields.Many2one('uom.uom', 'Capacity unit', domain=_get_capacity_filter)
    userdefine_ids = fields.One2many('circles.supplierinfo.userdefine', 'supplierinfo_id', string='User define')
    vendor_person_in_charge = fields.Many2one(string="PIC", related='name.user_id')
    vendor_tags = fields.Many2many(string="Source", related='name.category_id')
    vendor_evaluations = fields.Char('Evaluations', compute='_compute_evaluations')
    pricelist_ids = fields.One2many('product.pricelist.item', 'supplierinfo_id', string='Price list', search='_search_pricelist_ids')
    price_compute = fields.Char('Price', compute='_compute_pricelist')
    description = fields.Text('Product description')
    image = fields.Image("Product image")
    image_resize = fields.Image("Variant Image 128", related="image", max_width=128, max_height=128, store=True)
    state_id = fields.Many2one(strng="State",related='name.state_id')

    show_more_information_button = fields.Boolean("More information", default=False)

    vendor_criteria_search = fields.Many2one('circles.vendor.criteria', string='Have Criteria')
    incoterm_search = fields.Many2one('account.incoterms', string='Have Delivery type')

    def _search_pricelist_ids(self, operator, value):
        pass

    def _compute_evaluations(self):
        for record in self:
            record.vendor_evaluations = ' | '.join([ '%s (%s %s)'%(evaluation.criteria_id.name, evaluation.data, evaluation.rating) for evaluation in record.name.vendor_evaluation_ids])

    def _compute_pricelist(self):
        for record in self:
            record.price_compute = ' | '.join([ '%s %s %s'%(price_id.fixed_price, price_id.currency_id.name, price_id.incoterm_id.name if price_id.incoterm_id.name else " " ) for price_id in record.pricelist_ids ])


class ProductPriceListItem(models.Model):
    _inherit = 'product.pricelist.item'

    supplierinfo_id = fields.Many2one('product.supplierinfo')
    price_note = fields.Char('Price note', default='')
    logistic_type = fields.Selection([('f', 'Factory price'), ('b', 'FOB price')], string='Price type', default='f')
    incoterm_id = fields.Many2one('account.incoterms', string='Delivery type')
    max_quantity = fields.Integer('Max.Quantity')


class StockReport(models.TransientModel):
    _name = "wizard.stock.history"
    _description = "Current Stock History"

    def export_xls(self):
        ids = self.env.context.get('active_ids') # Get the ids from the current context

        data = {'ids': ids}
        return {
            'type': 'ir_actions_xlsx_download',
            'data': {'model': 'wizard.stock.history',
                     'options': json.dumps(data, default=date_utils.json_default),
                     'output_format': 'xlsx',
                     'report_name': 'Full vendors information',
                     }
        }

    def add_image(self, raw, coor, ws):
        binaryData = raw
        if not binaryData == False:
            data = base64.b64decode(binaryData)
            im = PILImage.open(io.BytesIO(data))
            img = Image(im)

            img.height = 128
            img.width = 128

            # set the width of the column 
            ws.column_dimensions[coor[0]].width = 35

            # set the height of the row 
            ws.row_dimensions[int(coor[1:])].height = 83

            ws.add_image(img, coor)


    def get_xlsx_report(self, data, response):
        wb = Workbook()
        with NamedTemporaryFile() as tmp:
            ws = wb.active

            # Start lines
            vendors = self.env['res.partner'].browse(data['ids'])

            ws.append(['Name', 'Address', 'Vat', 'PIC', 'Description', 'Phone', 'Mobile', 'Email', 'Web', 'Industry', 'Number empl', 'Comment', 'Geo', 'Contacts', 'Tags', 'Products', 'Product_img', 'Evaluation', 'Logs', 'Image Vendor'])
            line = 2
            for vendor in vendors:
                ws.cell(line, 1).value = vendor.name
                ws.cell(line, 2).value = vendor.street
                ws.cell(line, 3).value = vendor.vat
                ws.cell(line, 4).value = vendor.user_id.name
                ws.cell(line, 5).value = vendor.description
                ws.cell(line, 6).value = vendor.phone
                ws.cell(line, 7).value = vendor.mobile
                ws.cell(line, 8).value = vendor.email
                ws.cell(line, 9).value = vendor.website
                ws.cell(line, 10).value = vendor.industry_id.name
                ws.cell(line, 11).value = vendor.num_empl
                ws.cell(line, 12).value = vendor.comment
                ws.cell(line, 13).value = ", ".join([str(vendor.partner_latitude), str(vendor.partner_longitude)])

                max_line = max([len(vendor.message_ids), len(vendor.category_id), len(vendor.product_ids), len(vendor.vendor_evaluation_ids), len(vendor.child_ids)])
                print(max_line)

                temp_line = line
                for child in vendor.child_ids:
                    ws.cell(temp_line, 14).value = ', '.join([str(child.name), str(child.phone)])
                    temp_line +=1

                temp_line = line
                for tag in vendor.category_id:
                    ws.cell(temp_line, 15).value = tag.display_name
                    temp_line +=1


                temp_line = line
                for product in vendor.product_ids:
                    ws.cell(temp_line, 16).value = product.product_name
                    self.add_image(product.image_resize, "Q%d"%(temp_line), ws)
                    temp_line +=1

                temp_line = line
                for evaluation in vendor.vendor_evaluation_ids:
                    ws.cell(temp_line, 18).value = evaluation.criteria_id.name
                    temp_line +=1

                temp_line = line
                for log in vendor.message_ids:
                    ws.cell(temp_line, 19).value = log.body
                    temp_line +=1

                self.add_image(vendor.image_1920, "T%d"%(line), ws)

                line += max_line
            # End lines

            wb.save(tmp.name)
            output = io.BytesIO(tmp.read())


        output.seek(0)
        response.stream.write(output.read())
        output.close()
